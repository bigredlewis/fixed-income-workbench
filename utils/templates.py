"""
Excel template generators for file upload features.
Each function returns a BytesIO object ready for Streamlit download.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers


HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=11)


def _style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def create_credit_analysis_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Data"

    headers = [
        "Period", "Revenue ($M)", "EBITDA ($M)", "D&A ($M)",
        "Total Debt ($M)", "Cash & Equivalents ($M)", "Interest Expense ($M)",
        "Capital Expenditures ($M)", "Cash Taxes ($M)",
        "Working Capital Change ($M)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    # Sample data
    sample_data = [
        ["Q1 2024", 1250, 310, 45, 2100, 350, 55, 80, 40, 15],
        ["Q2 2024", 1300, 330, 46, 2050, 380, 54, 85, 42, -10],
        ["Q3 2024", 1280, 315, 46, 2000, 400, 53, 75, 38, 20],
        ["Q4 2024", 1350, 345, 47, 1950, 430, 52, 90, 45, -5],
    ]
    for row in sample_data:
        ws.append(row)

    # Add note
    note_row = len(sample_data) + 3
    ws.cell(row=note_row, column=1, value="Notes:").font = BOLD_FONT
    ws.cell(row=note_row + 1, column=1,
            value="- Working Capital Change: Positive = cash used, Negative = cash generated")
    ws.cell(row=note_row + 2, column=1,
            value="- Replace sample data with your issuer's actual financials")
    ws.cell(row=note_row + 3, column=1,
            value="- Period format: Q1 2024, Q2 2024, ... or FY 2024")

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_consensus_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consensus Comparison"

    headers = ["Metric", "Your Estimate", "Consensus"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    metrics = [
        "Revenue ($M)", "EBITDA ($M)", "EBITDA Margin (%)",
        "Net Income ($M)", "EPS ($)", "Total Debt ($M)",
        "Leverage (Debt/EBITDA)", "Interest Coverage (x)",
        "Free Cash Flow ($M)", "Capex ($M)", "Revenue Growth (%)",
    ]
    for m in metrics:
        ws.append([m, None, None])

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_portfolio_holdings_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    headers = [
        "Issuer", "CUSIP", "Bond Description", "Sector", "Credit Rating",
        "Coupon (%)", "Yield to Maturity (%)", "Duration (years)",
        "Spread/OAS (bps)", "Par Amount ($000s)", "Market Value ($000s)",
        "Weight (%)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    sample = [
        ["Apple Inc", "037833DX7", "3.25% 2029", "Technology", "AA+",
         3.25, 4.10, 4.2, 55, 500, 485, 20.0],
        ["JPMorgan Chase", "46647PBY0", "4.50% 2027", "Financials", "A-",
         4.50, 4.80, 2.8, 95, 600, 590, 24.0],
        ["Duke Energy", "26441CBJ4", "5.00% 2031", "Utilities", "BBB+",
         5.00, 5.15, 5.5, 120, 400, 395, 16.0],
        ["ExxonMobil", "30231GBL5", "4.75% 2030", "Energy", "AA-",
         4.75, 4.95, 4.8, 75, 500, 492, 20.0],
        ["HCA Healthcare", "404121AJ4", "5.50% 2028", "Healthcare", "BB+",
         5.50, 5.85, 3.1, 185, 500, 488, 20.0],
    ]
    for row in sample:
        ws.append(row)

    # Sectors note
    note_row = len(sample) + 3
    ws.cell(row=note_row, column=1, value="Valid Sectors:").font = BOLD_FONT
    sectors = [
        "Financials", "Industrials", "Utilities", "Energy", "Technology",
        "Healthcare", "Consumer Staples", "Consumer Discretionary",
        "Telecom/Media", "Real Estate", "Other",
    ]
    ws.cell(row=note_row + 1, column=1, value=", ".join(sectors))

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_three_statement_template() -> BytesIO:
    wb = Workbook()

    years = ["", "FY 2022", "FY 2023", "FY 2024", "FY 2025E"]

    # --- Income Statement ---
    ws_is = wb.active
    ws_is.title = "Income Statement"

    is_rows = [
        years,
        ["Revenue"],
        ["COGS"],
        ["Gross Profit"],
        ["SG&A"],
        ["R&D"],
        ["Other Operating Expenses"],
        ["EBITDA"],
        ["Depreciation & Amortization"],
        ["EBIT"],
        ["Interest Expense"],
        ["Pre-Tax Income"],
        ["Taxes"],
        ["Net Income"],
    ]
    for row in is_rows:
        ws_is.append(row)
    _style_header_row(ws_is, len(years))
    for row_idx in [1]:
        for col in range(1, len(years) + 1):
            ws_is.cell(row=row_idx, column=col).fill = HEADER_FILL
            ws_is.cell(row=row_idx, column=col).font = HEADER_FONT
    # Bold row labels
    for r in range(2, ws_is.max_row + 1):
        ws_is.cell(row=r, column=1).font = BOLD_FONT
    _auto_width(ws_is)

    # --- Balance Sheet ---
    ws_bs = wb.create_sheet("Balance Sheet")

    bs_rows = [
        years,
        ["Cash & Equivalents"],
        ["Accounts Receivable"],
        ["Inventory"],
        ["Other Current Assets"],
        ["Total Current Assets"],
        [""],
        ["PP&E (Net)"],
        ["Goodwill & Intangibles"],
        ["Other Long-Term Assets"],
        ["Total Assets"],
        [""],
        ["Accounts Payable"],
        ["Short-Term Debt"],
        ["Other Current Liabilities"],
        ["Total Current Liabilities"],
        [""],
        ["Long-Term Debt"],
        ["Other LT Liabilities"],
        ["Total Liabilities"],
        [""],
        ["Total Equity"],
        ["Total Liabilities & Equity"],
    ]
    for row in bs_rows:
        ws_bs.append(row)
    _style_header_row(ws_bs, len(years))
    for r in range(2, ws_bs.max_row + 1):
        ws_bs.cell(row=r, column=1).font = BOLD_FONT
    _auto_width(ws_bs)

    # --- Cash Flow Statement ---
    ws_cf = wb.create_sheet("Cash Flow Statement")

    cf_rows = [
        years,
        ["Net Income"],
        ["Depreciation & Amortization"],
        ["Stock-Based Compensation"],
        ["Other Non-Cash Items"],
        ["Changes in Working Capital"],
        ["Cash from Operations"],
        [""],
        ["Capital Expenditures"],
        ["Acquisitions"],
        ["Other Investing Activities"],
        ["Cash from Investing"],
        [""],
        ["Debt Issuance"],
        ["Debt Repayment"],
        ["Dividends"],
        ["Share Repurchases"],
        ["Other Financing Activities"],
        ["Cash from Financing"],
        [""],
        ["Net Change in Cash"],
    ]
    for row in cf_rows:
        ws_cf.append(row)
    _style_header_row(ws_cf, len(years))
    for r in range(2, ws_cf.max_row + 1):
        ws_cf.cell(row=r, column=1).font = BOLD_FONT
    _auto_width(ws_cf)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
